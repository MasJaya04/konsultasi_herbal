import csv
import io

from django.contrib import messages
from django.db import transaction
from django.db.models import ProtectedError
from django.forms import ValidationError
from django.http import HttpResponse
from django.shortcuts import redirect
from django.utils.text import slugify
from django.urls import reverse_lazy
from django.views.generic import CreateView, DeleteView, ListView, UpdateView
from openpyxl import Workbook, load_workbook

from accounts.mixins import AdminRequiredMixin

from .forms import ProductCategoryForm, ProductForm, ProductImportForm
from .models import Product, ProductCategory


def _build_query_string(request):
    params = request.GET.copy()
    params.pop("page", None)
    return params.urlencode()


class ProductCategoryListView(AdminRequiredMixin, ListView):
    model = ProductCategory
    template_name = "catalog/category_list.html"
    context_object_name = "categories"
    paginate_by = 10

    def get_queryset(self):
        queryset = ProductCategory.objects.order_by("name")
        search = self.request.GET.get("q", "").strip()
        if search:
            queryset = queryset.filter(name__icontains=search)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["search_value"] = self.request.GET.get("q", "").strip()
        context["query_string"] = _build_query_string(self.request)
        return context


class ProductCategoryCreateView(AdminRequiredMixin, CreateView):
    model = ProductCategory
    form_class = ProductCategoryForm
    template_name = "catalog/category_form.html"
    success_url = reverse_lazy("catalog:category_list")


class ProductCategoryUpdateView(AdminRequiredMixin, UpdateView):
    model = ProductCategory
    form_class = ProductCategoryForm
    template_name = "catalog/category_form.html"
    success_url = reverse_lazy("catalog:category_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Ubah Kategori Produk"
        context["submit_label"] = "Perbarui"
        return context


class ProductCategoryDeleteView(AdminRequiredMixin, DeleteView):
    model = ProductCategory
    template_name = "catalog/category_confirm_delete.html"
    success_url = reverse_lazy("catalog:category_list")

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        try:
            self.object.delete()
            messages.success(request, "Kategori produk berhasil dihapus.")
        except ProtectedError:
            messages.error(request, "Kategori produk masih dipakai oleh produk lain dan tidak bisa dihapus.")
        return redirect(self.success_url)


class ProductListView(AdminRequiredMixin, ListView):
    model = Product
    template_name = "catalog/product_list.html"
    context_object_name = "products"
    paginate_by = 10

    def post(self, request, *args, **kwargs):
        form = ProductImportForm(request.POST, request.FILES)
        if not form.is_valid():
            messages.error(request, "File CSV produk tidak valid.")
            return redirect("catalog:product_list")
        try:
            created_count, updated_count = self._import_csv(form.cleaned_data["csv_file"])
        except ValidationError as exc:
            messages.error(request, exc.message)
            return redirect("catalog:product_list")
        messages.success(
            request,
            f"Impor produk selesai. {created_count} data baru ditambahkan dan {updated_count} data diperbarui.",
        )
        return redirect("catalog:product_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["import_form"] = ProductImportForm()
        context["search_value"] = self.request.GET.get("q", "").strip()
        context["category_value"] = self.request.GET.get("category", "").strip()
        context["product_counts"] = {
            "total": Product.objects.count(),
            "active": Product.objects.filter(is_active=True).count(),
            "category_total": ProductCategory.objects.count(),
        }
        context["categories"] = ProductCategory.objects.order_by("name")
        context["query_string"] = _build_query_string(self.request)
        return context

    def get_queryset(self):
        queryset = Product.objects.select_related("category").order_by("name")
        search = self.request.GET.get("q", "").strip()
        category = self.request.GET.get("category", "").strip()
        if search:
            queryset = queryset.filter(name__icontains=search)
        if category:
            queryset = queryset.filter(category_id=category)
        return queryset

    @transaction.atomic
    def _import_csv(self, uploaded_file):
        rows = self._read_import_rows(uploaded_file)
        required_columns = {
            "category",
            "name",
            "slug",
            "description",
            "benefits",
            "usage_instructions",
            "contraindications",
            "is_active",
        }
        if not rows["fieldnames"]:
            raise ValidationError("Header CSV produk tidak ditemukan.")
        missing_columns = required_columns.difference({field.strip() for field in rows["fieldnames"] if field})
        if missing_columns:
            raise ValidationError(f"Header CSV produk kurang: {', '.join(sorted(missing_columns))}.")
        created_count = 0
        updated_count = 0
        for index, row in enumerate(rows["items"], start=2):
            category_name = (row.get("category") or "").strip()
            product_name = (row.get("name") or "").strip()
            if not category_name or not product_name:
                raise ValidationError(f"Baris {index} wajib memiliki category dan name.")
            category, _ = ProductCategory.objects.get_or_create(name=category_name)
            slug_value = (row.get("slug") or "").strip() or slugify(product_name)
            defaults = {
                "category": category,
                "name": product_name,
                "description": (row.get("description") or "").strip(),
                "benefits": (row.get("benefits") or "").strip(),
                "usage_instructions": (row.get("usage_instructions") or "").strip(),
                "contraindications": (row.get("contraindications") or "").strip(),
                "is_active": (row.get("is_active") or "").strip().lower() not in {"0", "false", "no", "tidak"},
            }
            if not defaults["description"] or not defaults["benefits"] or not defaults["usage_instructions"]:
                raise ValidationError(f"Baris {index} wajib memiliki description, benefits, dan usage_instructions.")
            _, created = Product.objects.update_or_create(slug=slug_value, defaults=defaults)
            if created:
                created_count += 1
            else:
                updated_count += 1
        return created_count, updated_count

    def _read_import_rows(self, uploaded_file):
        filename = uploaded_file.name.lower()
        if filename.endswith(".csv"):
            try:
                decoded = uploaded_file.read().decode("utf-8-sig")
            except UnicodeDecodeError as exc:
                raise ValidationError("File CSV produk harus menggunakan encoding UTF-8.") from exc
            reader = csv.DictReader(io.StringIO(decoded))
            return {"fieldnames": reader.fieldnames or [], "items": list(reader)}
        if filename.endswith(".xlsx"):
            workbook = load_workbook(uploaded_file, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return {"fieldnames": [], "items": []}
            fieldnames = [str(value).strip() if value is not None else "" for value in rows[0]]
            items = []
            for values in rows[1:]:
                if not any(value not in (None, "") for value in values):
                    continue
                items.append(
                    {
                        fieldnames[index]: "" if value is None else str(value)
                        for index, value in enumerate(values)
                        if index < len(fieldnames)
                    }
                )
            return {"fieldnames": fieldnames, "items": items}
        raise ValidationError("Gunakan file .csv atau .xlsx untuk import produk.")


class ProductCreateView(AdminRequiredMixin, CreateView):
    model = Product
    form_class = ProductForm
    template_name = "catalog/product_form.html"
    success_url = reverse_lazy("catalog:product_list")


class ProductUpdateView(AdminRequiredMixin, UpdateView):
    model = Product
    form_class = ProductForm
    template_name = "catalog/product_form.html"
    success_url = reverse_lazy("catalog:product_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Ubah Produk Herbal"
        context["submit_label"] = "Perbarui"
        return context


class ProductDeleteView(AdminRequiredMixin, DeleteView):
    model = Product
    template_name = "catalog/product_confirm_delete.html"
    success_url = reverse_lazy("catalog:product_list")

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        try:
            self.object.delete()
            messages.success(request, "Produk herbal berhasil dihapus.")
        except ProtectedError:
            messages.error(request, "Produk herbal masih dipakai oleh data lain dan tidak bisa dihapus.")
        return redirect(self.success_url)


class ProductExportView(AdminRequiredMixin, ListView):
    model = Product

    def get(self, request, *args, **kwargs):
        queryset = Product.objects.select_related("category").order_by("name")
        search = request.GET.get("q", "").strip()
        category = request.GET.get("category", "").strip()
        export_format = request.GET.get("format", "csv").strip().lower()
        if search:
            queryset = queryset.filter(name__icontains=search)
        if category:
            queryset = queryset.filter(category_id=category)
        rows = [
            [
                product.category.name,
                product.name,
                product.slug,
                product.description,
                product.benefits,
                product.usage_instructions,
                product.contraindications,
                "1" if product.is_active else "0",
            ]
            for product in queryset
        ]
        headers = ["category", "name", "slug", "description", "benefits", "usage_instructions", "contraindications", "is_active"]
        if export_format == "xlsx":
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(headers)
            for row in rows:
                sheet.append(row)
            buffer = io.BytesIO()
            workbook.save(buffer)
            response = HttpResponse(
                buffer.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = 'attachment; filename="produk-herbal.xlsx"'
            return response
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="produk-herbal.csv"'
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerows(rows)
        return response
