import csv
import io

from django.contrib.auth.views import LoginView, LogoutView
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import ProtectedError
from django.db.models import Count, Q
from django.forms import ValidationError
from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.views.generic import CreateView, DeleteView, ListView, RedirectView, UpdateView
from openpyxl import Workbook, load_workbook

from .forms import LoginForm, UserForm, UserImportForm, UserUpdateForm
from .mixins import AdminRequiredMixin
from .models import User
from consultations.models import create_activity_log


def _build_query_string(request):
    params = request.GET.copy()
    params.pop("page", None)
    return params.urlencode()


def _redirect_with_query(request, route_name):
    query_string = _build_query_string(request)
    url = str(reverse_lazy(route_name))
    if query_string:
        return redirect(f"{url}?{query_string}")
    return redirect(url)


def _log_activity(request, action, target_type, target_id="", description=""):
    create_activity_log(
        actor=getattr(request, "user", None) if getattr(request, "user", None) and request.user.is_authenticated else None,
        action=action,
        target_type=target_type,
        target_id=target_id,
        description=description,
        request_id=getattr(request, "request_id", ""),
    )


class HomeRedirectView(RedirectView):
    def get_redirect_url(self, *args, **kwargs):
        if not self.request.user.is_authenticated:
            return reverse_lazy("accounts:login")
        if self.request.user.role == "customer":
            return reverse_lazy("consultations:chat")
        return reverse_lazy("dashboard:home")


class AppLoginView(LoginView):
    form_class = LoginForm
    template_name = "accounts/login.html"

    def get_success_url(self):
        if self.request.user.role == "customer":
            return reverse_lazy("consultations:chat")
        return reverse_lazy("dashboard:home")


class AppLogoutView(LogoutView):
    pass


class UserListView(AdminRequiredMixin, ListView):
    model = User
    template_name = "accounts/user_list.html"
    context_object_name = "users"
    ordering = ("role", "username")
    paginate_by = 10

    def post(self, request, *args, **kwargs):
        if request.POST.get("form_type") == "bulk_action":
            selected_ids = request.POST.getlist("selected_ids")
            bulk_action = request.POST.get("bulk_action", "").strip()
            if not selected_ids:
                messages.error(request, "Pilih minimal satu pengguna untuk aksi massal.")
                return _redirect_with_query(request, "accounts:user_list")
            queryset = User.objects.filter(pk__in=selected_ids)
            if bulk_action == "delete":
                if queryset.filter(pk=request.user.pk).exists():
                    messages.error(request, "Akun yang sedang digunakan tidak bisa dihapus melalui aksi massal.")
                    return _redirect_with_query(request, "accounts:user_list")
                deleted_count = queryset.count()
                queryset.delete()
                messages.success(request, f"{deleted_count} pengguna berhasil dihapus.")
                _log_activity(request, "user_bulk_delete", "user", ",".join(selected_ids), f"Hapus massal {deleted_count} pengguna")
                return _redirect_with_query(request, "accounts:user_list")
            if bulk_action in {User.Role.ADMIN, User.Role.AI_TRAINER, User.Role.CUSTOMER}:
                updated_count = queryset.update(role=bulk_action)
                messages.success(request, f"{updated_count} pengguna berhasil diperbarui perannya.")
                _log_activity(request, "user_bulk_update", "user", ",".join(selected_ids), f"Perbarui massal peran menjadi {bulk_action} untuk {updated_count} pengguna")
                return _redirect_with_query(request, "accounts:user_list")
            messages.error(request, "Aksi massal pengguna tidak valid.")
            return _redirect_with_query(request, "accounts:user_list")
        form = UserImportForm(request.POST, request.FILES)
        if not form.is_valid():
            messages.error(request, "File impor pengguna tidak valid.")
            return _redirect_with_query(request, "accounts:user_list")
        try:
            created_count, updated_count = self._import_users(form.cleaned_data["import_file"])
        except ValidationError as exc:
            messages.error(request, exc.message)
            return _redirect_with_query(request, "accounts:user_list")
        messages.success(
            request,
            f"Impor pengguna selesai. {created_count} data baru ditambahkan dan {updated_count} data diperbarui.",
        )
        _log_activity(request, "user_import", "user", "", f"Impor pengguna {created_count} baru, {updated_count} diperbarui")
        return _redirect_with_query(request, "accounts:user_list")

    def get_queryset(self):
        queryset = User.objects.order_by("role", "username")
        role = self.request.GET.get("role", "").strip()
        search = self.request.GET.get("q", "").strip()
        if role:
            queryset = queryset.filter(role=role)
        if search:
            queryset = queryset.filter(
                Q(username__icontains=search)
                | Q(first_name__icontains=search)
                | Q(last_name__icontains=search)
                | Q(email__icontains=search)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["filter_values"] = {
            "role": self.request.GET.get("role", "").strip(),
            "q": self.request.GET.get("q", "").strip(),
        }
        context["role_counts"] = {
            "admin": User.objects.filter(role=User.Role.ADMIN).count(),
            "ai_trainer": User.objects.filter(role=User.Role.AI_TRAINER).count(),
            "customer": User.objects.filter(role=User.Role.CUSTOMER).count(),
        }
        context["import_form"] = UserImportForm()
        context["query_string"] = _build_query_string(self.request)
        return context

    def _import_users(self, uploaded_file):
        rows = self._read_import_rows(uploaded_file)
        required_columns = {"first_name", "last_name", "username", "email", "role", "password"}
        if not rows["fieldnames"]:
            raise ValidationError("Header file user tidak ditemukan.")
        missing_columns = required_columns.difference({field.strip() for field in rows["fieldnames"] if field})
        if missing_columns:
            raise ValidationError(f"Header import user kurang: {', '.join(sorted(missing_columns))}.")
        valid_roles = {choice[0] for choice in User.Role.choices}
        created_count = 0
        updated_count = 0
        for index, row in enumerate(rows["items"], start=2):
            username = (row.get("username") or "").strip()
            role = (row.get("role") or "").strip()
            password = (row.get("password") or "").strip()
            if not username or not role:
                raise ValidationError(f"Baris {index} wajib memiliki nama pengguna dan peran.")
            if role not in valid_roles:
                raise ValidationError(f"Baris {index} memiliki peran tidak valid: {role}.")
            user, created = User.objects.update_or_create(
                username=username,
                defaults={
                    "first_name": (row.get("first_name") or "").strip(),
                    "last_name": (row.get("last_name") or "").strip(),
                    "email": (row.get("email") or "").strip(),
                    "role": role,
                },
            )
            if password:
                if len(password) < 8:
                    raise ValidationError(f"Kata sandi pada baris {index} minimal 8 karakter.")
                user.set_password(password)
                user.save(update_fields=["password"])
            if created:
                created_count += 1
            else:
                updated_count += 1
        return created_count, updated_count

    def _read_import_rows(self, uploaded_file):
        filename = uploaded_file.name.lower()
        if filename.endswith(".csv"):
            try:
                decoded = uploaded_file.read().decode("utf-8-sig")
            except UnicodeDecodeError as exc:
                raise ValidationError("File CSV user harus menggunakan encoding UTF-8.") from exc
            reader = csv.DictReader(io.StringIO(decoded))
            return {"fieldnames": reader.fieldnames or [], "items": list(reader)}
        if filename.endswith(".xlsx"):
            workbook = load_workbook(uploaded_file, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return {"fieldnames": [], "items": []}
            fieldnames = [str(value).strip() if value is not None else "" for value in rows[0]]
            items = []
            for values in rows[1:]:
                if not any(value not in (None, "") for value in values):
                    continue
                items.append(
                    {
                        fieldnames[index]: "" if value is None else str(value)
                        for index, value in enumerate(values)
                        if index < len(fieldnames)
                    }
                )
            return {"fieldnames": fieldnames, "items": items}
        raise ValidationError("Gunakan file .csv atau .xlsx untuk impor pengguna.")


class UserCreateView(AdminRequiredMixin, CreateView):
    model = User
    form_class = UserForm
    template_name = "accounts/user_form.html"
    success_url = reverse_lazy("accounts:user_list")

    def form_valid(self, form):
        response = super().form_valid(form)
        _log_activity(self.request, "user_create", "user", self.object.pk, f"Membuat pengguna {self.object.username}")
        return response


class UserUpdateView(AdminRequiredMixin, UpdateView):
    model = User
    form_class = UserUpdateForm
    template_name = "accounts/user_form.html"
    success_url = reverse_lazy("accounts:user_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Ubah Pengguna"
        context["submit_label"] = "Perbarui"
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        _log_activity(self.request, "user_update", "user", self.object.pk, f"Memperbarui pengguna {self.object.username}")
        return response


class UserDeleteView(AdminRequiredMixin, DeleteView):
    model = User
    template_name = "accounts/user_confirm_delete.html"
    success_url = reverse_lazy("accounts:user_list")

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        if self.object == request.user:
            messages.error(request, "Akun yang sedang digunakan tidak bisa dihapus.")
            return redirect(self.success_url)
        try:
            self.object.delete()
            messages.success(request, "Pengguna berhasil dihapus.")
            _log_activity(request, "user_delete", "user", self.object.pk, f"Menghapus user {self.object.username}")
        except ProtectedError:
            messages.error(request, "Pengguna masih dipakai oleh data lain dan tidak bisa dihapus.")
        return redirect(self.success_url)


class UserExportView(AdminRequiredMixin, ListView):
    model = User

    def get(self, request, *args, **kwargs):
        queryset = User.objects.order_by("role", "username")
        role = request.GET.get("role", "").strip()
        search = request.GET.get("q", "").strip()
        export_format = request.GET.get("format", "csv").strip().lower()
        if role:
            queryset = queryset.filter(role=role)
        if search:
            queryset = queryset.filter(
                Q(username__icontains=search)
                | Q(first_name__icontains=search)
                | Q(last_name__icontains=search)
                | Q(email__icontains=search)
            )
        headers = ["first_name", "last_name", "username", "email", "role"]
        rows = [[user.first_name, user.last_name, user.username, user.email, user.role] for user in queryset]
        if export_format == "xlsx":
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(headers)
            for row in rows:
                sheet.append(row)
            buffer = io.BytesIO()
            workbook.save(buffer)
            response = HttpResponse(
                buffer.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = 'attachment; filename="users.xlsx"'
            _log_activity(request, "user_export", "user", "xlsx", "Export user format xlsx")
            return response
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="users.csv"'
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerows(rows)
        _log_activity(request, "user_export", "user", "csv", "Export user format csv")
        return response
