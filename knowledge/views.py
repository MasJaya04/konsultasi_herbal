import csv
import io

from django.contrib import messages
from django.db import transaction
from django.db.models import Count, Q
from django.db.models import ProtectedError
from django.forms import ValidationError
from django.http import HttpResponse
from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.utils import timezone
from django.views.generic import CreateView, DeleteView, ListView, UpdateView
from openpyxl import Workbook, load_workbook

from accounts.mixins import StaffRequiredMixin
from catalog.models import Product
from consultations.models import AIResponseReview, UnansweredQuestion

from .forms import KnowledgeCategoryForm, KnowledgeEntryForm, KnowledgeImportForm
from .models import KnowledgeCategory, KnowledgeEntry, _entry_matches_unanswered, refresh_knowledge_quality


def _build_query_string(request):
    params = request.GET.copy()
    params.pop("page", None)
    return params.urlencode()


class KnowledgeCategoryListView(StaffRequiredMixin, ListView):
    model = KnowledgeCategory
    template_name = "knowledge/category_list.html"
    context_object_name = "categories"
    paginate_by = 10

    def get_queryset(self):
        queryset = KnowledgeCategory.objects.order_by("name")
        search = self.request.GET.get("q", "").strip()
        if search:
            queryset = queryset.filter(name__icontains=search)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["search_value"] = self.request.GET.get("q", "").strip()
        context["query_string"] = _build_query_string(self.request)
        return context


class KnowledgeCategoryCreateView(StaffRequiredMixin, CreateView):
    model = KnowledgeCategory
    form_class = KnowledgeCategoryForm
    template_name = "knowledge/category_form.html"
    success_url = reverse_lazy("knowledge:category_list")


class KnowledgeCategoryUpdateView(StaffRequiredMixin, UpdateView):
    model = KnowledgeCategory
    form_class = KnowledgeCategoryForm
    template_name = "knowledge/category_form.html"
    success_url = reverse_lazy("knowledge:category_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Ubah Kategori Pengetahuan"
        context["submit_label"] = "Perbarui"
        return context


class KnowledgeCategoryDeleteView(StaffRequiredMixin, DeleteView):
    model = KnowledgeCategory
    template_name = "knowledge/category_confirm_delete.html"
    success_url = reverse_lazy("knowledge:category_list")

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        try:
            self.object.delete()
            messages.success(request, "Kategori pengetahuan berhasil dihapus.")
        except ProtectedError:
            messages.error(request, "Kategori pengetahuan masih dipakai oleh entri lain dan tidak bisa dihapus.")
        return redirect(self.success_url)


class KnowledgeEntryListView(StaffRequiredMixin, ListView):
    model = KnowledgeEntry
    template_name = "knowledge/entry_list.html"
    context_object_name = "entries"
    paginate_by = 10

    def post(self, request, *args, **kwargs):
        form = KnowledgeImportForm(request.POST, request.FILES)
        if not form.is_valid():
            messages.error(request, "File CSV basis pengetahuan tidak valid.")
            return redirect("knowledge:entry_list")
        try:
            created_count, updated_count = self._import_csv(form.cleaned_data["csv_file"])
        except ValidationError as exc:
            messages.error(request, exc.message)
            return redirect("knowledge:entry_list")
        messages.success(
            request,
            f"Impor basis pengetahuan selesai. {created_count} entri baru ditambahkan dan {updated_count} entri diperbarui.",
        )
        return redirect("knowledge:entry_list")

    def get_queryset(self):
        queryset = KnowledgeEntry.objects.select_related("category", "created_by", "product")
        source_type = self.request.GET.get("source_type", "").strip()
        status = self.request.GET.get("status", "").strip()
        quality_status = self.request.GET.get("quality_status", "").strip()
        category = self.request.GET.get("category", "").strip()
        product = self.request.GET.get("product", "").strip()
        search = self.request.GET.get("q", "").strip()
        if source_type:
            queryset = queryset.filter(source_type=source_type)
        if status:
            queryset = queryset.filter(status=status)
        if quality_status:
            queryset = queryset.filter(quality_status=quality_status)
        if category:
            queryset = queryset.filter(category_id=category)
        if product:
            queryset = queryset.filter(product_id=product)
        if search:
            queryset = queryset.filter(
                Q(title__icontains=search)
                | Q(question__icontains=search)
                | Q(answer__icontains=search)
                | Q(keywords__icontains=search)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["filter_values"] = {
            "source_type": self.request.GET.get("source_type", "").strip(),
            "status": self.request.GET.get("status", "").strip(),
            "quality_status": self.request.GET.get("quality_status", "").strip(),
            "category": self.request.GET.get("category", "").strip(),
            "product": self.request.GET.get("product", "").strip(),
            "q": self.request.GET.get("q", "").strip(),
        }
        context["knowledge_counts"] = {
            "total": KnowledgeEntry.objects.count(),
            "published": KnowledgeEntry.objects.filter(status=KnowledgeEntry.Status.PUBLISHED).count(),
            "draft": KnowledgeEntry.objects.filter(status=KnowledgeEntry.Status.DRAFT).count(),
        }
        context["quality_counts"] = {
            "strong": KnowledgeEntry.objects.filter(quality_status=KnowledgeEntry.QualityStatus.STRONG).count(),
            "needs_review": KnowledgeEntry.objects.filter(quality_status=KnowledgeEntry.QualityStatus.NEEDS_REVIEW).count(),
            "weak": KnowledgeEntry.objects.filter(quality_status=KnowledgeEntry.QualityStatus.WEAK).count(),
        }
        product_overview = []
        unanswered_counts = dict(
            UnansweredQuestion.objects.filter(product__isnull=False)
            .values("product_id")
            .annotate(total=Count("id"))
            .values_list("product_id", "total")
        )
        for product_item in Product.objects.filter(is_active=True).order_by("name"):
            product_entries = KnowledgeEntry.objects.filter(product=product_item)
            total_entries = product_entries.count()
            strong_entries = product_entries.filter(quality_status=KnowledgeEntry.QualityStatus.STRONG).count()
            weak_entries = product_entries.filter(quality_status=KnowledgeEntry.QualityStatus.WEAK).count()
            needs_review_entries = product_entries.filter(quality_status=KnowledgeEntry.QualityStatus.NEEDS_REVIEW).count()
            if weak_entries > 0 or unanswered_counts.get(product_item.id, 0) > 0:
                health_label = "Perlu tindakan"
                health_class = "bg-rose-50 text-rose-700"
            elif needs_review_entries > 0:
                health_label = "Perlu review"
                health_class = "bg-amber-50 text-amber-700"
            else:
                health_label = "Stabil"
                health_class = "bg-emerald-50 text-emerald-700"
            product_overview.append(
                {
                    "product": product_item,
                    "total_entries": total_entries,
                    "strong_entries": strong_entries,
                    "weak_entries": weak_entries,
                    "needs_review_entries": needs_review_entries,
                    "unanswered_total": unanswered_counts.get(product_item.id, 0),
                    "health_label": health_label,
                    "health_class": health_class,
                }
            )
        context["product_overview"] = product_overview
        context["import_form"] = KnowledgeImportForm()
        context["categories"] = KnowledgeCategory.objects.order_by("name")
        context["products"] = Product.objects.filter(is_active=True).order_by("name")
        context["query_string"] = _build_query_string(self.request)
        return context

    @transaction.atomic
    def _import_csv(self, uploaded_file):
        rows = self._read_import_rows(uploaded_file)
        required_columns = {"product", "category", "title", "question", "answer", "source_type", "keywords", "status"}
        if not rows["fieldnames"]:
            raise ValidationError("Header CSV basis pengetahuan tidak ditemukan.")
        missing_columns = required_columns.difference({field.strip() for field in rows["fieldnames"] if field})
        if missing_columns:
            raise ValidationError(f"Header CSV basis pengetahuan kurang: {', '.join(sorted(missing_columns))}.")
        valid_source_types = {choice[0] for choice in KnowledgeEntry.SourceType.choices}
        valid_statuses = {choice[0] for choice in KnowledgeEntry.Status.choices}
        created_count = 0
        updated_count = 0
        for index, row in enumerate(rows["items"], start=2):
            category_name = (row.get("category") or "").strip()
            product_name = (row.get("product") or "").strip()
            title = (row.get("title") or "").strip()
            answer = (row.get("answer") or "").strip()
            source_type = (row.get("source_type") or "").strip().lower()
            status = (row.get("status") or "").strip().lower()
            if not category_name or not title or not answer:
                raise ValidationError(f"Baris {index} wajib memiliki kategori, judul, dan jawaban.")
            if source_type not in valid_source_types:
                raise ValidationError(f"Baris {index} memiliki tipe sumber tidak valid: {source_type}.")
            if status not in valid_statuses:
                raise ValidationError(f"Baris {index} memiliki status tidak valid: {status}.")
            product = None
            if product_name:
                product = Product.objects.filter(name=product_name, is_active=True).first()
            if product_name and not product:
                raise ValidationError(f"Baris {index} memiliki produk yang tidak ditemukan: {product_name}.")
            category, _ = KnowledgeCategory.objects.get_or_create(name=category_name)
            _, created = KnowledgeEntry.objects.update_or_create(
                product=product,
                title=title,
                defaults={
                    "category": category,
                    "question": (row.get("question") or "").strip(),
                    "answer": answer,
                    "source_type": source_type,
                    "keywords": (row.get("keywords") or "").strip(),
                    "status": status,
                    "quality_status": KnowledgeEntry.QualityStatus.NEEDS_REVIEW,
                    "created_by": self.request.user,
                },
            )
            if created:
                created_count += 1
            else:
                updated_count += 1
        return created_count, updated_count

    def _read_import_rows(self, uploaded_file):
        filename = uploaded_file.name.lower()
        if filename.endswith(".csv"):
            try:
                decoded = uploaded_file.read().decode("utf-8-sig")
            except UnicodeDecodeError as exc:
                raise ValidationError("File CSV basis pengetahuan harus menggunakan encoding UTF-8.") from exc
            reader = csv.DictReader(io.StringIO(decoded))
            return {"fieldnames": reader.fieldnames or [], "items": list(reader)}
        if filename.endswith(".xlsx"):
            workbook = load_workbook(uploaded_file, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return {"fieldnames": [], "items": []}
            fieldnames = [str(value).strip() if value is not None else "" for value in rows[0]]
            items = []
            for values in rows[1:]:
                if not any(value not in (None, "") for value in values):
                    continue
                items.append(
                    {
                        fieldnames[index]: "" if value is None else str(value)
                        for index, value in enumerate(values)
                        if index < len(fieldnames)
                    }
                )
            return {"fieldnames": fieldnames, "items": items}
        raise ValidationError("Gunakan file .csv atau .xlsx untuk impor basis pengetahuan.")


class KnowledgeQualityMonitorView(StaffRequiredMixin, ListView):
    model = KnowledgeEntry
    template_name = "knowledge/quality_monitor.html"
    context_object_name = "entries"
    paginate_by = 12

    def post(self, request, *args, **kwargs):
        entry = KnowledgeEntry.objects.filter(pk=request.POST.get("entry_id")).first()
        quality_status = request.POST.get("quality_status", "").strip()
        next_url = request.POST.get("next", "").strip()
        valid_statuses = {choice[0] for choice in KnowledgeEntry.QualityStatus.choices}
        if not entry or quality_status not in valid_statuses:
            messages.error(request, "Aksi kualitas pengetahuan tidak valid.")
            return redirect(next_url or "knowledge:quality_monitor")
        entry.quality_status = quality_status
        entry.reviewed_by = request.user
        entry.reviewed_at = timezone.now()
        if quality_status == KnowledgeEntry.QualityStatus.STRONG:
            entry.review_note = "Ditandai kuat oleh pelatih AI dari monitor kualitas."
        elif quality_status == KnowledgeEntry.QualityStatus.WEAK:
            entry.review_note = "Ditandai lemah oleh pelatih AI dari monitor kualitas dan perlu revisi pengetahuan."
        else:
            entry.review_note = "Ditandai perlu ditinjau oleh pelatih AI dari monitor kualitas."
        entry.save(update_fields=["quality_status", "review_note", "reviewed_by", "reviewed_at", "updated_at"])
        messages.success(request, f"Kualitas pengetahuan '{entry.title}' berhasil diperbarui.")
        return redirect(next_url or "knowledge:quality_monitor")

    def get_queryset(self):
        queryset = KnowledgeEntry.objects.select_related("product", "category", "reviewed_by").filter(product__isnull=False)
        product = self.request.GET.get("product", "").strip()
        quality_status = self.request.GET.get("quality_status", "").strip()
        signal = self.request.GET.get("signal", "").strip()
        search = self.request.GET.get("q", "").strip()
        if product:
            queryset = queryset.filter(product_id=product)
        if quality_status:
            queryset = queryset.filter(quality_status=quality_status)
        if search:
            queryset = queryset.filter(
                Q(title__icontains=search)
                | Q(question__icontains=search)
                | Q(answer__icontains=search)
                | Q(keywords__icontains=search)
            )
        entry_ids = list(queryset.values_list("id", flat=True))
        review_summary = {
            item["message__knowledge_entries"]: {
                "accurate": item["accurate"],
                "needs_revision": item["needs_revision"],
                "incorrect": item["incorrect"],
            }
            for item in AIResponseReview.objects.filter(message__knowledge_entries__in=entry_ids)
            .values("message__knowledge_entries")
            .annotate(
                accurate=Count("id", filter=Q(verdict=AIResponseReview.Verdict.ACCURATE)),
                needs_revision=Count("id", filter=Q(verdict=AIResponseReview.Verdict.NEEDS_REVISION)),
                incorrect=Count("id", filter=Q(verdict=AIResponseReview.Verdict.INCORRECT)),
            )
        }
        unanswered_map = {}
        unanswered_by_product = {}
        for item in UnansweredQuestion.objects.filter(
            product_id__in=queryset.values_list("product_id", flat=True),
            status=UnansweredQuestion.Status.OPEN,
        ).select_related("product"):
            unanswered_by_product.setdefault(item.product_id, []).append(item)
        matched_entry_ids = []
        for entry in queryset:
            product_unanswered = unanswered_by_product.get(entry.product_id, [])
            unanswered_total = sum(1 for unanswered in product_unanswered if _entry_matches_unanswered(entry, unanswered))
            unanswered_map[entry.pk] = unanswered_total
            review_data = review_summary.get(entry.pk, {"accurate": 0, "needs_revision": 0, "incorrect": 0})
            if signal == "incorrect" and review_data["incorrect"] == 0:
                continue
            if signal == "needs_revision" and review_data["needs_revision"] == 0:
                continue
            if signal == "unanswered" and unanswered_total == 0:
                continue
            matched_entry_ids.append(entry.pk)
        filtered_queryset = queryset.filter(pk__in=matched_entry_ids) if signal else queryset
        self.review_summary = review_summary
        self.unanswered_map = unanswered_map
        return filtered_queryset.order_by("product__name", "quality_status", "title")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        entries = list(context["entries"])
        alert_count = 0
        for entry in entries:
            review_data = self.review_summary.get(entry.pk, {"accurate": 0, "needs_revision": 0, "incorrect": 0})
            unanswered_total = self.unanswered_map.get(entry.pk, 0)
            reason_parts = []
            if review_data["incorrect"]:
                reason_parts.append(f"{review_data['incorrect']} incorrect")
            if review_data["needs_revision"]:
                reason_parts.append(f"{review_data['needs_revision']} needs revision")
            if unanswered_total:
                reason_parts.append(f"{unanswered_total} unanswered terkait")
            if entry.quality_status != KnowledgeEntry.QualityStatus.STRONG or reason_parts:
                alert_count += 1
            entry.monitor_stats = {
                "accurate": review_data["accurate"],
                "needs_revision": review_data["needs_revision"],
                "incorrect": review_data["incorrect"],
                "unanswered": unanswered_total,
                "reason_text": " | ".join(reason_parts) if reason_parts else "Belum ada sinyal risiko baru.",
            }
        context["entries"] = entries
        context["filter_values"] = {
            "product": self.request.GET.get("product", "").strip(),
            "quality_status": self.request.GET.get("quality_status", "").strip(),
            "signal": self.request.GET.get("signal", "").strip(),
            "q": self.request.GET.get("q", "").strip(),
        }
        context["products"] = Product.objects.filter(is_active=True).order_by("name")
        context["monitor_summary"] = {
            "total": KnowledgeEntry.objects.count(),
            "attention": KnowledgeEntry.objects.filter(
                quality_status__in=[KnowledgeEntry.QualityStatus.NEEDS_REVIEW, KnowledgeEntry.QualityStatus.WEAK],
            ).count(),
            "weak": KnowledgeEntry.objects.filter(quality_status=KnowledgeEntry.QualityStatus.WEAK).count(),
            "alerts_on_page": alert_count,
        }
        context["query_string"] = _build_query_string(self.request)
        return context

class KnowledgeEntryCreateView(StaffRequiredMixin, CreateView):
    model = KnowledgeEntry
    form_class = KnowledgeEntryForm
    template_name = "knowledge/entry_form.html"
    success_url = reverse_lazy("knowledge:entry_list")

    def get_initial(self):
        initial = super().get_initial()
        title = self.request.GET.get("title")
        question = self.request.GET.get("question")
        keywords = self.request.GET.get("keywords")
        answer = self.request.GET.get("answer")
        product = self.request.GET.get("product")
        if title:
            initial["title"] = title
        if question:
            initial["question"] = question
        if keywords:
            initial["keywords"] = keywords
        if answer:
            initial["answer"] = answer
        if product:
            initial["product"] = product
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        unanswered_id = self.request.GET.get("unanswered_id")
        if unanswered_id:
            context["unanswered_item"] = UnansweredQuestion.objects.select_related("customer", "product").filter(pk=unanswered_id).first()
        return context

    @transaction.atomic
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.reviewed_by = self.request.user
        form.instance.reviewed_at = timezone.now()
        response = super().form_valid(form)
        unanswered_id = self.request.GET.get("unanswered_id")
        if unanswered_id:
            unanswered = UnansweredQuestion.objects.filter(pk=unanswered_id).first()
            if unanswered:
                unanswered.status = UnansweredQuestion.Status.RESOLVED
                unanswered.resolved_entry = self.object
                unanswered.save(update_fields=["status", "resolved_entry", "updated_at"])
        refresh_knowledge_quality(product_ids=[self.object.product_id], entry_ids=[self.object.pk])
        return response


class KnowledgeEntryUpdateView(StaffRequiredMixin, UpdateView):
    model = KnowledgeEntry
    form_class = KnowledgeEntryForm
    template_name = "knowledge/entry_form.html"
    success_url = reverse_lazy("knowledge:entry_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Ubah Entri Pengetahuan"
        context["submit_label"] = "Perbarui"
        return context

    def form_valid(self, form):
        form.instance.reviewed_by = self.request.user
        form.instance.reviewed_at = timezone.now()
        response = super().form_valid(form)
        refresh_knowledge_quality(product_ids=[self.object.product_id], entry_ids=[self.object.pk])
        return response


class KnowledgeEntryDeleteView(StaffRequiredMixin, DeleteView):
    model = KnowledgeEntry
    template_name = "knowledge/entry_confirm_delete.html"
    success_url = reverse_lazy("knowledge:entry_list")

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        try:
            self.object.delete()
            messages.success(request, "Entri pengetahuan berhasil dihapus.")
        except ProtectedError:
            messages.error(request, "Entri pengetahuan masih dipakai oleh data lain dan tidak bisa dihapus.")
        return redirect(self.success_url)


class KnowledgeEntryExportView(StaffRequiredMixin, ListView):
    model = KnowledgeEntry

    def get(self, request, *args, **kwargs):
        queryset = KnowledgeEntry.objects.select_related("category", "created_by", "product").filter(product__isnull=False).order_by("title")
        source_type = request.GET.get("source_type", "").strip()
        status = request.GET.get("status", "").strip()
        quality_status = request.GET.get("quality_status", "").strip()
        category = request.GET.get("category", "").strip()
        product = request.GET.get("product", "").strip()
        search = request.GET.get("q", "").strip()
        export_format = request.GET.get("format", "csv").strip().lower()
        if source_type:
            queryset = queryset.filter(source_type=source_type)
        if status:
            queryset = queryset.filter(status=status)
        if quality_status:
            queryset = queryset.filter(quality_status=quality_status)
        if category:
            queryset = queryset.filter(category_id=category)
        if product:
            queryset = queryset.filter(product_id=product)
        if search:
            queryset = queryset.filter(
                Q(title__icontains=search)
                | Q(question__icontains=search)
                | Q(answer__icontains=search)
                | Q(keywords__icontains=search)
            )
        headers = ["product", "category", "title", "question", "answer", "source_type", "keywords", "status", "quality_status", "review_note"]
        rows = [
            [entry.product.name if entry.product else "", entry.category.name, entry.title, entry.question, entry.answer, entry.source_type, entry.keywords, entry.status, entry.quality_status, entry.review_note]
            for entry in queryset
        ]
        if export_format == "xlsx":
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(headers)
            for row in rows:
                sheet.append(row)
            buffer = io.BytesIO()
            workbook.save(buffer)
            response = HttpResponse(
                buffer.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = 'attachment; filename="knowledge-base.xlsx"'
            return response
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="knowledge-base.csv"'
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerows(rows)
        return response
