from django.conf import settings
from django.contrib import messages
from django.db import transaction
from django.db.models import Count, Max, Q
from django.db.models.functions import TruncDate
from django.forms import ValidationError
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.urls import reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views import View
from django.views.generic import CreateView, DeleteView, DetailView, FormView, ListView, TemplateView, UpdateView
from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
import io
import csv

from accounts.mixins import AdminRequiredMixin, NonCustomerRequiredMixin, StaffRequiredMixin
from catalog.models import Product, ProductCategory
from knowledge.models import KnowledgeEntry, refresh_knowledge_quality

from .forms import AIResponseReviewForm, ConsultationPromptForm, ReviewImportForm, UnansweredImportForm, UnansweredQuestionForm
from .models import AIResponseReview, ActivityLog, ConsultationMessage, ConsultationSession, UnansweredQuestion, create_activity_log
from .services import analyze_prompt, generate_consultation_answer


def _build_query_string(request):
    params = request.GET.copy()
    params.pop("page", None)
    return params.urlencode()


def _redirect_with_query(request, route_name):
    query_string = _build_query_string(request)
    url = str(reverse_lazy(route_name))
    if query_string:
        return redirect(f"{url}?{query_string}")
    return redirect(url)


def _read_import_rows(uploaded_file, error_label):
    filename = uploaded_file.name.lower()
    if filename.endswith(".csv"):
        try:
            decoded = uploaded_file.read().decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise ValidationError(f"File {error_label} harus menggunakan encoding UTF-8.") from exc
        reader = csv.DictReader(io.StringIO(decoded))
        return {"fieldnames": reader.fieldnames or [], "items": list(reader)}
    if filename.endswith(".xlsx"):
        workbook = load_workbook(uploaded_file, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return {"fieldnames": [], "items": []}
        fieldnames = [str(value).strip() if value is not None else "" for value in rows[0]]
        items = []
        for values in rows[1:]:
            if not any(value not in (None, "") for value in values):
                continue
            items.append(
                {
                    fieldnames[index]: "" if value is None else str(value)
                    for index, value in enumerate(values)
                    if index < len(fieldnames)
                }
            )
        return {"fieldnames": fieldnames, "items": items}
    raise ValidationError(f"Gunakan file .csv atau .xlsx untuk {error_label}.")


def _log_activity(request, action, target_type, target_id="", description=""):
    create_activity_log(
        actor=getattr(request, "user", None) if getattr(request, "user", None) and request.user.is_authenticated else None,
        action=action,
        target_type=target_type,
        target_id=target_id,
        description=description,
        request_id=getattr(request, "request_id", ""),
    )


def _refresh_quality_from_messages(messages_queryset):
    messages = list(messages_queryset.prefetch_related("knowledge_entries").select_related("product"))
    product_ids = {message.product_id for message in messages if message.product_id}
    entry_ids = set()
    for message in messages:
        entry_ids.update(message.knowledge_entries.values_list("id", flat=True))
    if product_ids or entry_ids:
        refresh_knowledge_quality(product_ids=list(product_ids), entry_ids=list(entry_ids))


def _refresh_quality_from_unanswered(unanswered_queryset):
    items = list(unanswered_queryset.select_related("product", "resolved_entry"))
    product_ids = {item.product_id for item in items if item.product_id}
    entry_ids = {item.resolved_entry_id for item in items if item.resolved_entry_id}
    if product_ids or entry_ids:
        refresh_knowledge_quality(product_ids=list(product_ids), entry_ids=list(entry_ids))


class DashboardHomeView(StaffRequiredMixin, TemplateView):
    template_name = "dashboard/home.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        filter_start = parse_date(self.request.GET.get("start_date", "").strip())
        filter_end = parse_date(self.request.GET.get("end_date", "").strip())
        report_role = self.request.GET.get("report_role", "").strip()
        report_category = self.request.GET.get("report_category", "").strip()
        report_product = self.request.GET.get("report_product", "").strip()
        today = timezone.localdate()
        trend_end = filter_end or today
        trend_start = filter_start or (trend_end - timezone.timedelta(days=6))
        session_queryset = ConsultationSession.objects.all()
        unanswered_queryset = UnansweredQuestion.objects.all()
        ai_message_queryset = ConsultationMessage.objects.filter(sender=ConsultationMessage.Sender.AI)
        review_queryset = AIResponseReview.objects.all()
        knowledge_queryset = KnowledgeEntry.objects.all()
        if filter_start:
            session_queryset = session_queryset.filter(created_at__date__gte=filter_start)
            unanswered_queryset = unanswered_queryset.filter(created_at__date__gte=filter_start)
            ai_message_queryset = ai_message_queryset.filter(created_at__date__gte=filter_start)
            review_queryset = review_queryset.filter(created_at__date__gte=filter_start)
            knowledge_queryset = knowledge_queryset.filter(created_at__date__gte=filter_start)
        if filter_end:
            session_queryset = session_queryset.filter(created_at__date__lte=filter_end)
            unanswered_queryset = unanswered_queryset.filter(created_at__date__lte=filter_end)
            ai_message_queryset = ai_message_queryset.filter(created_at__date__lte=filter_end)
            review_queryset = review_queryset.filter(created_at__date__lte=filter_end)
            knowledge_queryset = knowledge_queryset.filter(created_at__date__lte=filter_end)
        filtered_user_queryset = self.request.user.__class__.objects.all()
        filtered_session_queryset = session_queryset
        filtered_unanswered_queryset = unanswered_queryset
        filtered_product_queryset = Product.objects.all()
        if report_role:
            filtered_user_queryset = filtered_user_queryset.filter(role=report_role)
            filtered_session_queryset = filtered_session_queryset.filter(user__role=report_role)
            filtered_unanswered_queryset = filtered_unanswered_queryset.filter(customer__role=report_role)
        if report_category:
            filtered_product_queryset = filtered_product_queryset.filter(category_id=report_category)
        if report_product:
            filtered_product_queryset = filtered_product_queryset.filter(pk=report_product)
            session_queryset = session_queryset.filter(Q(product_id=report_product) | Q(messages__product_id=report_product)).distinct()
            unanswered_queryset = unanswered_queryset.filter(product_id=report_product)
            ai_message_queryset = ai_message_queryset.filter(product_id=report_product)
            review_queryset = review_queryset.filter(message__product_id=report_product)
            knowledge_queryset = knowledge_queryset.filter(product_id=report_product)
            filtered_session_queryset = filtered_session_queryset.filter(Q(product_id=report_product) | Q(messages__product_id=report_product)).distinct()
            filtered_unanswered_queryset = filtered_unanswered_queryset.filter(product_id=report_product)
        total_reviews = review_queryset.count()
        accurate_reviews = review_queryset.filter(verdict=AIResponseReview.Verdict.ACCURATE).count()
        needs_revision_reviews = review_queryset.filter(verdict=AIResponseReview.Verdict.NEEDS_REVISION).count()
        incorrect_reviews = review_queryset.filter(verdict=AIResponseReview.Verdict.INCORRECT).count()
        pending_reviews = ai_message_queryset.filter(
            sender=ConsultationMessage.Sender.AI,
            reviews__isnull=True,
        ).count()
        fallback_answers = ai_message_queryset.filter(
            sender=ConsultationMessage.Sender.AI,
            response_state=ConsultationMessage.ResponseState.FALLBACK,
        ).count()
        low_confidence_total = ai_message_queryset.filter(
            sender=ConsultationMessage.Sender.AI,
            confidence_score__isnull=False,
            confidence_score__lt=0.70,
        ).count()
        total_ai_messages = ai_message_queryset.count()
        unanswered_open = unanswered_queryset.filter(status=UnansweredQuestion.Status.OPEN).count()
        published_knowledge = knowledge_queryset.filter(status=KnowledgeEntry.Status.PUBLISHED).count()
        unresolved_ratio = round((unanswered_open / published_knowledge) * 100, 1) if published_knowledge else 0
        review_completion = round((total_reviews / total_ai_messages) * 100, 1) if total_ai_messages else 0
        answer_success_rate = round(((total_ai_messages - fallback_answers) / total_ai_messages) * 100, 1) if total_ai_messages else 0
        quality_score = round(
            max(
                0,
                min(
                    100,
                    55
                    + (accurate_reviews * 12)
                    - (needs_revision_reviews * 7)
                    - (incorrect_reviews * 12)
                    - (fallback_answers * 4)
                    - (low_confidence_total * 2),
                ),
            ),
            1,
        )
        context["metrics"] = {
            "users": self.request.user.__class__.objects.count(),
            "products": Product.objects.count(),
            "knowledge_entries": knowledge_queryset.count(),
            "consultations": session_queryset.count(),
            "unanswered_questions": unanswered_open,
        }
        context["role_metrics"] = {
            "admin": self.request.user.__class__.objects.filter(role="admin").count(),
            "ai_trainer": self.request.user.__class__.objects.filter(role="ai_trainer").count(),
            "customer": self.request.user.__class__.objects.filter(role="customer").count(),
        }
        context["latest_unanswered"] = unanswered_queryset.select_related("customer")[:5]
        context["top_sessions"] = session_queryset.annotate(message_count=Count("messages")).select_related("user")[:5]
        context["trainer_metrics"] = {
            "pending_reviews": pending_reviews,
            "fallback_answers": fallback_answers,
        }
        context["top_fallback_questions"] = (
            unanswered_queryset.values("question")
            .annotate(total=Count("id"), latest_at=Max("created_at"))
            .order_by("-total", "-latest_at")[:5]
        )
        context["top_knowledge_usage"] = (
            knowledge_queryset.annotate(usage_count=Count("consultation_messages"))
            .filter(usage_count__gt=0)
            .order_by("-usage_count", "title")[:5]
        )
        context["review_summary"] = {
            "accurate": accurate_reviews,
            "needs_revision": needs_revision_reviews,
            "incorrect": incorrect_reviews,
        }
        context["low_confidence_messages"] = (
            ai_message_queryset.filter(
                sender=ConsultationMessage.Sender.AI,
                confidence_score__isnull=False,
                confidence_score__lt=0.70,
            )
            .select_related("session__user")
            .order_by("confidence_score", "-created_at")[:5]
        )
        context["system_status"] = {
            "database": "PostgreSQL" if settings.USE_POSTGRES and settings.DATABASES["default"]["ENGINE"].endswith("postgresql") else "SQLite",
            "ollama_model": settings.OLLAMA_MODEL,
            "ollama_url": settings.OLLAMA_BASE_URL,
            "rate_limit": settings.RATE_LIMIT_PER_MINUTE,
        }
        context["operational_summary"] = {
            "quality_score": quality_score,
            "review_completion": review_completion,
            "answer_success_rate": answer_success_rate,
            "unresolved_ratio": unresolved_ratio,
            "pending_reviews": pending_reviews,
            "fallback_answers": fallback_answers,
            "low_confidence_total": low_confidence_total,
            "published_knowledge": published_knowledge,
        }
        reviewed_total = accurate_reviews + needs_revision_reviews + incorrect_reviews
        context["dashboard_charts"] = {
            "review_segments": [
                {
                    "label": "Akurat",
                    "value": accurate_reviews,
                    "width": round((accurate_reviews / reviewed_total) * 100, 1) if reviewed_total else 0,
                    "bar_class": "bg-emerald-500",
                    "text_class": "text-emerald-700",
                    "bg_class": "bg-emerald-50",
                },
                {
                    "label": "Perlu Revisi",
                    "value": needs_revision_reviews,
                    "width": round((needs_revision_reviews / reviewed_total) * 100, 1) if reviewed_total else 0,
                    "bar_class": "bg-amber-500",
                    "text_class": "text-amber-700",
                    "bg_class": "bg-amber-50",
                },
                {
                    "label": "Tidak Tepat",
                    "value": incorrect_reviews,
                    "width": round((incorrect_reviews / reviewed_total) * 100, 1) if reviewed_total else 0,
                    "bar_class": "bg-rose-500",
                    "text_class": "text-rose-700",
                    "bg_class": "bg-rose-50",
                },
            ],
            "pipeline_segments": [
                {
                    "label": "Terjawab",
                    "value": max(total_ai_messages - fallback_answers, 0),
                    "width": answer_success_rate,
                    "bar_class": "bg-sky-500",
                    "text_class": "text-sky-700",
                    "bg_class": "bg-sky-50",
                },
                {
                    "label": "Cadangan",
                    "value": fallback_answers,
                    "width": round((fallback_answers / total_ai_messages) * 100, 1) if total_ai_messages else 0,
                    "bar_class": "bg-rose-500",
                    "text_class": "text-rose-700",
                    "bg_class": "bg-rose-50",
                },
                {
                    "label": "Menunggu Evaluasi",
                    "value": pending_reviews,
                    "width": round((pending_reviews / total_ai_messages) * 100, 1) if total_ai_messages else 0,
                    "bar_class": "bg-amber-500",
                    "text_class": "text-amber-700",
                    "bg_class": "bg-amber-50",
                },
            ],
        }
        context["priority_actions"] = [
            {
                "title": "Tinjau pertanyaan belum terjawab",
                "value": unanswered_open,
                "description": "Pertanyaan pengguna yang masih menunggu basis pengetahuan atau penanganan pelatih.",
                "url": reverse_lazy("consultations:unanswered_list"),
                "tone": "rose",
            },
            {
                "title": "Evaluasi jawaban AI tertunda",
                "value": pending_reviews,
                "description": "Respons AI yang belum dievaluasi kualitasnya oleh pelatih AI.",
                "url": reverse_lazy("consultations:review_list"),
                "tone": "amber",
            },
            {
                "title": "Pengetahuan siap diperkaya",
                "value": published_knowledge,
                "description": "Total entri pengetahuan aktif yang menjadi dasar jawaban RAG saat ini.",
                "url": reverse_lazy("knowledge:entry_list"),
                "tone": "emerald",
            },
        ]
        consultation_trend_map = {
            item["day"]: item["total"]
            for item in ConsultationSession.objects.filter(created_at__date__gte=trend_start, created_at__date__lte=trend_end)
            .annotate(day=TruncDate("created_at"))
            .values("day")
            .annotate(total=Count("id"))
        }
        unanswered_trend_map = {
            item["day"]: item["total"]
            for item in UnansweredQuestion.objects.filter(created_at__date__gte=trend_start, created_at__date__lte=trend_end)
            .annotate(day=TruncDate("created_at"))
            .values("day")
            .annotate(total=Count("id"))
        }
        trend_items = []
        max_trend_value = 1
        for offset in range(7):
            day = trend_start + timezone.timedelta(days=offset)
            consultation_total = consultation_trend_map.get(day, 0)
            unanswered_total = unanswered_trend_map.get(day, 0)
            max_trend_value = max(max_trend_value, consultation_total, unanswered_total)
            trend_items.append(
                {
                    "label": day.strftime("%d %b"),
                    "consultations": consultation_total,
                    "unanswered": unanswered_total,
                }
            )
        for item in trend_items:
            item["consultation_height"] = max(12, round((item["consultations"] / max_trend_value) * 100)) if item["consultations"] else 12
            item["unanswered_height"] = max(12, round((item["unanswered"] / max_trend_value) * 100)) if item["unanswered"] else 12
        context["daily_trends"] = trend_items
        context["dashboard_filters"] = {
            "start_date": self.request.GET.get("start_date", "").strip(),
            "end_date": self.request.GET.get("end_date", "").strip(),
            "report_role": report_role,
            "report_category": report_category,
            "report_product": report_product,
        }
        context["report_categories"] = ProductCategory.objects.order_by("name")
        context["report_all_products"] = Product.objects.select_related("category").order_by("name")
        context["report_summary"] = {
            "users": filtered_user_queryset.count(),
            "sessions": filtered_session_queryset.count(),
            "unanswered": filtered_unanswered_queryset.count(),
            "products": filtered_product_queryset.count(),
        }
        context["report_products"] = filtered_product_queryset.select_related("category").order_by("name")[:5]
        selected_report_product = filtered_product_queryset.select_related("category").first() if report_product else None
        context["selected_report_product"] = selected_report_product
        if selected_report_product:
            product_knowledge_queryset = KnowledgeEntry.objects.filter(product=selected_report_product)
            product_session_queryset = ConsultationSession.objects.filter(
                Q(product=selected_report_product) | Q(messages__product=selected_report_product)
            ).distinct()
            product_unanswered_queryset = UnansweredQuestion.objects.filter(product=selected_report_product)
            product_ai_message_queryset = ConsultationMessage.objects.filter(
                sender=ConsultationMessage.Sender.AI,
                product=selected_report_product,
            )
            product_review_queryset = AIResponseReview.objects.filter(message__product=selected_report_product)
            product_review_total = product_review_queryset.count()
            product_accuracy = round(
                (product_review_queryset.filter(verdict=AIResponseReview.Verdict.ACCURATE).count() / product_review_total) * 100,
                1,
            ) if product_review_total else 0
            product_fallback_answers = product_ai_message_queryset.filter(
                response_state=ConsultationMessage.ResponseState.FALLBACK,
            ).count()
            product_answer_total = product_ai_message_queryset.count()
            product_success_rate = round(
                ((product_answer_total - product_fallback_answers) / product_answer_total) * 100,
                1,
            ) if product_answer_total else 0
            context["product_focus_metrics"] = {
                "knowledge_entries": product_knowledge_queryset.count(),
                "consultations": product_session_queryset.count(),
                "unanswered_open": product_unanswered_queryset.filter(status=UnansweredQuestion.Status.OPEN).count(),
                "reviewed_messages": product_ai_message_queryset.filter(reviews__isnull=False).distinct().count(),
                "accuracy_rate": product_accuracy,
                "success_rate": product_success_rate,
            }
            context["product_focus_knowledge"] = product_knowledge_queryset.order_by("title")[:6]
        else:
            context["product_focus_metrics"] = None
            context["product_focus_knowledge"] = []
        context["latest_activity_logs"] = ActivityLog.objects.select_related("actor")[:6]
        return context


class DashboardExportView(StaffRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        export_format = request.GET.get("format", "xlsx").strip().lower()
        dashboard_view = DashboardHomeView()
        dashboard_view.request = request
        context = dashboard_view.get_context_data()
        category_lookup = dict(ProductCategory.objects.values_list("id", "name"))
        product_lookup = dict(Product.objects.values_list("id", "name"))
        category_label = category_lookup.get(int(context["dashboard_filters"]["report_category"]), "Semua kategori") if context["dashboard_filters"]["report_category"].isdigit() else "Semua kategori"
        product_label = product_lookup.get(int(context["dashboard_filters"]["report_product"]), "Semua produk") if context["dashboard_filters"]["report_product"].isdigit() else "Semua produk"
        if export_format == "pdf":
            buffer = io.BytesIO()
            pdf = canvas.Canvas(buffer, pagesize=A4)
            width, height = A4
            y = height - 50
            lines = [
                "Dasbor Konsultasi Herbal",
                f"Periode: {context['dashboard_filters']['start_date'] or '-'} s.d. {context['dashboard_filters']['end_date'] or '-'}",
                f"Saringan Peran: {context['dashboard_filters']['report_role'] or 'Semua peran'}",
                f"Saringan Kategori: {category_label}",
                f"Saringan Produk: {product_label}",
                "",
                f"Total Pengguna: {context['metrics']['users']}",
                f"Produk Herbal: {context['metrics']['products']}",
                f"Basis Pengetahuan: {context['metrics']['knowledge_entries']}",
                f"Sesi Konsultasi: {context['metrics']['consultations']}",
                f"Belum Terjawab: {context['metrics']['unanswered_questions']}",
                f"Skor Kualitas: {context['operational_summary']['quality_score']}",
                f"Jawaban Berhasil: {context['operational_summary']['answer_success_rate']}%",
                f"Evaluasi Selesai: {context['operational_summary']['review_completion']}%",
                f"Jawaban Cadangan: {context['operational_summary']['fallback_answers']}",
                f"Keyakinan Rendah: {context['operational_summary']['low_confidence_total']}",
                f"Menunggu Evaluasi: {context['operational_summary']['pending_reviews']}",
            ]
            for line in lines:
                pdf.drawString(40, y, str(line))
                y -= 18
            pdf.showPage()
            pdf.save()
            response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
            response["Content-Disposition"] = 'attachment; filename="dashboard-konsultasi-herbal.pdf"'
            _log_activity(request, "dashboard_export", "dashboard", "pdf", "Ekspor dasbor format pdf")
            return response
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Dasbor"
        rows = [
            ["Periode Mulai", context["dashboard_filters"]["start_date"] or "-"],
            ["Periode Selesai", context["dashboard_filters"]["end_date"] or "-"],
            ["Saringan Peran", context["dashboard_filters"]["report_role"] or "Semua peran"],
            ["Saringan Kategori", category_label],
            ["Saringan Produk", product_label],
            ["Total Pengguna", context["metrics"]["users"]],
            ["Produk Herbal", context["metrics"]["products"]],
            ["Basis Pengetahuan", context["metrics"]["knowledge_entries"]],
            ["Sesi Konsultasi", context["metrics"]["consultations"]],
            ["Belum Terjawab", context["metrics"]["unanswered_questions"]],
            ["Skor Kualitas", context["operational_summary"]["quality_score"]],
            ["Jawaban Berhasil (%)", context["operational_summary"]["answer_success_rate"]],
            ["Evaluasi Selesai (%)", context["operational_summary"]["review_completion"]],
            ["Jawaban Cadangan", context["operational_summary"]["fallback_answers"]],
            ["Keyakinan Rendah", context["operational_summary"]["low_confidence_total"]],
            ["Menunggu Evaluasi", context["operational_summary"]["pending_reviews"]],
        ]
        for row in rows:
            sheet.append(row)
        buffer = io.BytesIO()
        workbook.save(buffer)
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="dashboard-konsultasi-herbal.xlsx"'
        _log_activity(request, "dashboard_export", "dashboard", export_format, f"Ekspor dasbor format {export_format}")
        return response


class ActivityLogListView(AdminRequiredMixin, ListView):
    model = ActivityLog
    template_name = "consultations/activity_log_list.html"
    context_object_name = "logs"
    paginate_by = 15

    def get_queryset(self):
        queryset = ActivityLog.objects.select_related("actor")
        search = self.request.GET.get("q", "").strip()
        if search:
            queryset = queryset.filter(
                Q(action__icontains=search)
                | Q(target_type__icontains=search)
                | Q(description__icontains=search)
                | Q(actor__username__icontains=search)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["search_value"] = self.request.GET.get("q", "").strip()
        context["query_string"] = _build_query_string(self.request)
        return context


class ConsultationChatView(LoginRequiredMixin, FormView):
    template_name = "consultations/chat.html"
    form_class = ConsultationPromptForm
    success_url = reverse_lazy("consultations:chat")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        session = self._get_selected_session()
        current_product = session.latest_referenced_product or session.product
        initial = kwargs.get("initial", {})
        if current_product and "product" not in initial:
            initial["product"] = current_product
        kwargs["initial"] = initial
        return kwargs

    def post(self, request, *args, **kwargs):
        if "new_chat" in request.POST:
            session = self._create_new_session()
            return redirect(f"{reverse_lazy('consultations:chat')}?session={session.id}")
        return super().post(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["session"] = self._get_selected_session()
        context["chat_messages"] = context["session"].messages.select_related("product").prefetch_related("knowledge_entries")
        sessions = list(ConsultationSession.objects.filter(user=self.request.user).select_related("product").order_by("-updated_at"))
        context["sessions"] = sessions
        context["session_groups"] = self._build_session_groups(sessions)
        return context

    def form_valid(self, form):
        session = self._get_selected_session()
        prompt = form.cleaned_data["prompt"]
        selected_product = form.cleaned_data.get("product")
        analysis = analyze_prompt(
            prompt,
            active_product=session.latest_referenced_product,
            selected_product=selected_product,
        )
        product = analysis["product"]
        if analysis.get("direct_answer"):
            self.active_session = session
            result = {
                "answer": analysis["direct_answer"],
                "confidence": analysis["direct_confidence"],
                "used_rag": analysis["direct_used_rag"],
                "knowledge_entries": [],
                "source_summary": analysis["direct_source_summary"],
                "response_state": analysis["direct_response_state"],
                "reason": "",
                "store_unanswered": False,
            }
            return self._save_chat_result(session, product, prompt, result)
        if analysis["needs_clarification"]:
            self.active_session = session
            result = {
                "answer": analysis["clarification_message"],
                "confidence": 0,
                "used_rag": False,
                "knowledge_entries": [],
                "source_summary": f"OPTIONS:{'|'.join(analysis.get('clarification_options', []))}" if analysis.get("clarification_options") else "",
                "response_state": "answered",
                "reason": "",
                "store_unanswered": False,
            }
            return self._save_chat_result(session, product, prompt, result)
        if product and session.product_id != product.id:
            session.product = product
            session.save(update_fields=["product", "updated_at"])
        self.active_session = session
        should_update_title = not session.messages.exists() or session.title in {
            f"Sesi konsultasi {self.request.user.username}",
            "Percakapan baru",
        }
        try:
            result = generate_consultation_answer(analysis["processed_prompt"], product=product, request=self.request)
        except Exception:
            result = {
                "answer": "",
                "confidence": 0,
                "used_rag": True,
                "knowledge_entries": [],
                "source_summary": "",
                "response_state": "fallback",
                "reason": "Terjadi kendala saat menghubungi layanan AI.",
                "store_unanswered": False,
            }
        return self._save_chat_result(session, product, prompt, result, should_update_title=should_update_title)

    def _save_chat_result(self, session, product, prompt, result, should_update_title=True):
        with transaction.atomic():
            if should_update_title:
                session.title = self._build_session_title(prompt)
                session.save(update_fields=["title", "updated_at"])
            ConsultationMessage.objects.create(
                session=session,
                product=product,
                sender=ConsultationMessage.Sender.USER,
                content=prompt,
            )
            if result["answer"]:
                ai_message = ConsultationMessage.objects.create(
                    session=session,
                    product=product,
                    sender=ConsultationMessage.Sender.AI,
                    content=result["answer"],
                    confidence_score=result["confidence"],
                    used_rag=result["used_rag"],
                    response_state=result["response_state"],
                    source_summary=result["source_summary"],
                )
                if result["knowledge_entries"]:
                    ai_message.knowledge_entries.set(result["knowledge_entries"])
            elif result.get("store_unanswered", True):
                UnansweredQuestion.objects.create(
                    session=session,
                    customer=self.request.user,
                    product=product,
                    question=prompt,
                    reason=result["reason"],
                )
                if product:
                    refresh_knowledge_quality(product_ids=[product.id])
                ConsultationMessage.objects.create(
                    session=session,
                    product=product,
                    sender=ConsultationMessage.Sender.AI,
                    content=self._build_fallback_answer(result["reason"]),
                    confidence_score=0,
                    used_rag=result["used_rag"],
                    response_state=result["response_state"],
                    source_summary=result["source_summary"],
                )
            else:
                ConsultationMessage.objects.create(
                    session=session,
                    product=product,
                    sender=ConsultationMessage.Sender.AI,
                    content=self._build_fallback_answer(result["reason"]),
                    confidence_score=0,
                    used_rag=result["used_rag"],
                    response_state=result["response_state"],
                    source_summary=result["source_summary"],
                )
        return redirect(self.get_success_url())

    def get_success_url(self):
        session = getattr(self, "active_session", None) or self._get_selected_session()
        return f"{reverse_lazy('consultations:chat')}?session={session.id}"

    def _get_or_create_session(self):
        session, _ = ConsultationSession.objects.get_or_create(
            user=self.request.user,
            status=ConsultationSession.Status.ACTIVE,
            defaults={"title": "Percakapan baru"},
        )
        return session

    def _get_selected_session(self):
        session_id = self.request.GET.get("session")
        if session_id:
            session = ConsultationSession.objects.filter(user=self.request.user, pk=session_id).first()
            if session:
                return session
        return self._get_or_create_session()

    def _create_new_session(self, product=None):
        ConsultationSession.objects.filter(
            user=self.request.user,
            status=ConsultationSession.Status.ACTIVE,
        ).update(status=ConsultationSession.Status.CLOSED)
        return ConsultationSession.objects.create(
            user=self.request.user,
            product=product,
            title="Percakapan baru",
            status=ConsultationSession.Status.ACTIVE,
        )

    def _build_session_title(self, prompt):
        cleaned_prompt = " ".join(prompt.split())
        for prefix in [
            "saya ingin tanya",
            "saya mau tanya",
            "saya ingin bertanya",
            "saya mau bertanya",
            "ingin tanya",
            "mau tanya",
            "apakah",
            "halo",
            "hi",
        ]:
            lowered = cleaned_prompt.lower()
            if lowered.startswith(prefix):
                cleaned_prompt = cleaned_prompt[len(prefix):].strip(" ,.:;!?")
                break
        if not cleaned_prompt:
            cleaned_prompt = "Percakapan baru"
        if len(cleaned_prompt) <= 60:
            return cleaned_prompt[:1].upper() + cleaned_prompt[1:]
        return f"{cleaned_prompt[:57].rstrip()}..."

    def _build_fallback_answer(self, reason):
        if reason == "Produk herbal belum disebutkan di prompt.":
            return "Silakan sebutkan nama produk herbal yang ingin ditanyakan, misalnya: produk Detogreen untuk apa?"
        if reason == "Sistem tidak menemukan konteks yang relevan di basis pengetahuan.":
            return "Maaf, saya belum menemukan informasi yang cukup di basis pengetahuan untuk menjawab pertanyaan ini. Pertanyaan Anda sudah disimpan agar dapat ditinjau lebih lanjut oleh pelatih AI."
        if reason == "Terjadi kendala saat menghubungi layanan AI.":
            return "Maaf, saat ini terjadi kendala saat memproses permintaan Anda. Silakan coba beberapa saat lagi."
        return "Maaf, saya belum memiliki konteks yang cukup untuk menjawab pertanyaan ini. Pertanyaan Anda sudah disimpan untuk ditinjau lebih lanjut."

    def _build_session_groups(self, sessions):
        now = timezone.localtime()
        groups = [
            {"label": "Hari ini", "items": []},
            {"label": "Kemarin", "items": []},
            {"label": "7 hari terakhir", "items": []},
            {"label": "Lebih lama", "items": []},
        ]
        for item in sessions:
            updated_at = timezone.localtime(item.updated_at)
            days_difference = (now.date() - updated_at.date()).days
            if days_difference == 0:
                groups[0]["items"].append(item)
            elif days_difference == 1:
                groups[1]["items"].append(item)
            elif days_difference <= 7:
                groups[2]["items"].append(item)
            else:
                groups[3]["items"].append(item)
        return [group for group in groups if group["items"]]


class ConsultationSessionDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk, *args, **kwargs):
        session = get_object_or_404(ConsultationSession, pk=pk, user=request.user)
        session.delete()
        next_session = ConsultationSession.objects.filter(user=request.user).order_by("-updated_at").first()
        if next_session:
            return redirect(f"{reverse_lazy('consultations:chat')}?session={next_session.id}")
        new_session = ConsultationSession.objects.create(
            user=request.user,
            title="Percakapan baru",
            status=ConsultationSession.Status.ACTIVE,
        )
        return redirect(f"{reverse_lazy('consultations:chat')}?session={new_session.id}")


class ConsultationSessionDetailView(NonCustomerRequiredMixin, DetailView):
    model = ConsultationSession
    template_name = "consultations/session_detail.html"
    context_object_name = "session"

    def get_queryset(self):
        queryset = ConsultationSession.objects.prefetch_related("messages__knowledge_entries", "messages__product").select_related("user", "product")
        if self.request.user.role in {"admin", "ai_trainer"} or self.request.user.is_superuser:
            return queryset
        return queryset.filter(user=self.request.user)


class ConsultationHistoryView(NonCustomerRequiredMixin, ListView):
    template_name = "consultations/history.html"
    context_object_name = "sessions"
    paginate_by = 10

    def get_queryset(self):
        search = self.request.GET.get("q", "").strip()
        product = self.request.GET.get("product", "").strip()
        start_date = parse_date(self.request.GET.get("start_date", "").strip())
        end_date = parse_date(self.request.GET.get("end_date", "").strip())
        if self.request.user.role in {"admin", "ai_trainer"} or self.request.user.is_superuser:
            queryset = ConsultationSession.objects.select_related("user", "product").prefetch_related("messages__product")
        else:
            queryset = ConsultationSession.objects.filter(user=self.request.user).select_related("user", "product").prefetch_related("messages__product")
        if search:
            queryset = queryset.filter(Q(title__icontains=search) | Q(user__username__icontains=search))
        if product:
            queryset = queryset.filter(Q(product_id=product) | Q(messages__product_id=product))
        if start_date:
            queryset = queryset.filter(created_at__date__gte=start_date)
        if end_date:
            queryset = queryset.filter(created_at__date__lte=end_date)
        return queryset.distinct()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["search_value"] = self.request.GET.get("q", "").strip()
        context["product_value"] = self.request.GET.get("product", "").strip()
        context["start_date"] = self.request.GET.get("start_date", "").strip()
        context["end_date"] = self.request.GET.get("end_date", "").strip()
        context["products"] = Product.objects.filter(is_active=True).order_by("name")
        context["query_string"] = _build_query_string(self.request)
        return context


class UnansweredQuestionListView(StaffRequiredMixin, ListView):
    model = UnansweredQuestion
    template_name = "consultations/unanswered_list.html"
    context_object_name = "items"
    paginate_by = 10

    def post(self, request, *args, **kwargs):
        if request.POST.get("form_type") == "import":
            form = UnansweredImportForm(request.POST, request.FILES)
            if not form.is_valid():
                messages.error(request, "File impor pertanyaan belum terjawab tidak valid.")
                return _redirect_with_query(request, "consultations:unanswered_list")
            created_count, updated_count = self._import_unanswered(form.cleaned_data["import_file"])
            messages.success(request, f"Impor pertanyaan belum terjawab selesai. {created_count} data baru dan {updated_count} data diperbarui.")
            _log_activity(request, "unanswered_import", "unanswered_question", "", f"Impor pertanyaan belum terjawab {created_count} baru, {updated_count} diperbarui")
            return _redirect_with_query(request, "consultations:unanswered_list")
        if request.POST.get("form_type") == "bulk_action":
            selected_ids = request.POST.getlist("selected_ids")
            bulk_action = request.POST.get("bulk_action", "").strip()
            if not selected_ids:
                messages.error(request, "Pilih minimal satu pertanyaan belum terjawab untuk aksi massal.")
                return _redirect_with_query(request, "consultations:unanswered_list")
            queryset = UnansweredQuestion.objects.filter(pk__in=selected_ids)
            if bulk_action == "delete":
                deleted_count = queryset.count()
                product_ids = [item for item in queryset.values_list("product_id", flat=True) if item]
                entry_ids = [item for item in queryset.values_list("resolved_entry_id", flat=True) if item]
                queryset.delete()
                refresh_knowledge_quality(product_ids=product_ids, entry_ids=entry_ids)
                messages.success(request, f"{deleted_count} pertanyaan belum terjawab berhasil dihapus.")
                _log_activity(request, "unanswered_bulk_delete", "unanswered_question", ",".join(selected_ids), f"Hapus massal {deleted_count} pertanyaan belum terjawab")
                return _redirect_with_query(request, "consultations:unanswered_list")
            if bulk_action in {
                UnansweredQuestion.Status.OPEN,
                UnansweredQuestion.Status.REVIEWED,
                UnansweredQuestion.Status.RESOLVED,
            }:
                updated_count = queryset.update(status=bulk_action)
                _refresh_quality_from_unanswered(queryset)
                messages.success(request, f"{updated_count} pertanyaan belum terjawab berhasil diperbarui.")
                _log_activity(request, "unanswered_bulk_update", "unanswered_question", ",".join(selected_ids), f"Perbarui massal status menjadi {bulk_action} untuk {updated_count} pertanyaan belum terjawab")
                return _redirect_with_query(request, "consultations:unanswered_list")
            messages.error(request, "Aksi massal pertanyaan belum terjawab tidak valid.")
            return _redirect_with_query(request, "consultations:unanswered_list")
        return _redirect_with_query(request, "consultations:unanswered_list")

    def get_queryset(self):
        queryset = UnansweredQuestion.objects.select_related("customer", "resolved_entry", "session", "product")
        status_value = self.request.GET.get("status", "").strip()
        product_value = self.request.GET.get("product", "").strip()
        search = self.request.GET.get("q", "").strip()
        start_date = parse_date(self.request.GET.get("start_date", "").strip())
        end_date = parse_date(self.request.GET.get("end_date", "").strip())
        if product_value:
            queryset = queryset.filter(product_id=product_value)
        if status_value:
            queryset = queryset.filter(status=status_value)
        if search:
            queryset = queryset.filter(
                Q(question__icontains=search)
                | Q(reason__icontains=search)
                | Q(customer__username__icontains=search)
            )
        if start_date:
            queryset = queryset.filter(created_at__date__gte=start_date)
        if end_date:
            queryset = queryset.filter(created_at__date__lte=end_date)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["status_value"] = self.request.GET.get("status", "").strip()
        context["product_value"] = self.request.GET.get("product", "").strip()
        context["search_value"] = self.request.GET.get("q", "").strip()
        context["start_date"] = self.request.GET.get("start_date", "").strip()
        context["end_date"] = self.request.GET.get("end_date", "").strip()
        product_value = self.request.GET.get("product", "").strip()
        base_queryset = UnansweredQuestion.objects.all()
        if product_value:
            base_queryset = base_queryset.filter(product_id=product_value)
        context["status_counts"] = {
            "open": base_queryset.filter(status=UnansweredQuestion.Status.OPEN).count(),
            "reviewed": base_queryset.filter(status=UnansweredQuestion.Status.REVIEWED).count(),
            "resolved": base_queryset.filter(status=UnansweredQuestion.Status.RESOLVED).count(),
        }
        context["trainer_summary"] = {
            "total": base_queryset.count(),
            "with_resolution": base_queryset.filter(resolved_entry__isnull=False).count(),
            "without_resolution": base_queryset.filter(resolved_entry__isnull=True).count(),
        }
        context["products"] = Product.objects.filter(is_active=True).order_by("name")
        context["import_form"] = UnansweredImportForm()
        context["query_string"] = _build_query_string(self.request)
        return context

    def _import_unanswered(self, uploaded_file):
        rows = _read_import_rows(uploaded_file, "impor pertanyaan belum terjawab")
        required_columns = {"session_id", "customer_username", "product", "question", "reason", "status"}
        if not rows["fieldnames"]:
            raise ValidationError("Header file pertanyaan belum terjawab tidak ditemukan.")
        missing_columns = required_columns.difference({field.strip() for field in rows["fieldnames"] if field})
        if missing_columns:
            raise ValidationError(f"Header impor pertanyaan belum terjawab kurang: {', '.join(sorted(missing_columns))}.")
        created_count = 0
        updated_count = 0
        affected_product_ids = set()
        for row in rows["items"]:
            session = ConsultationSession.objects.filter(pk=(row.get("session_id") or "").strip()).first()
            customer = self.request.user.__class__.objects.filter(username=(row.get("customer_username") or "").strip()).first()
            product = Product.objects.filter(name=(row.get("product") or "").strip(), is_active=True).first()
            if not session or not customer or not product:
                continue
            affected_product_ids.add(product.id)
            defaults = {
                "product": product,
                "reason": (row.get("reason") or "").strip(),
                "status": (row.get("status") or "").strip() or UnansweredQuestion.Status.OPEN,
            }
            _, created = UnansweredQuestion.objects.update_or_create(
                session=session,
                customer=customer,
                question=(row.get("question") or "").strip(),
                defaults=defaults,
            )
            if created:
                created_count += 1
            else:
                updated_count += 1
        if affected_product_ids:
            refresh_knowledge_quality(product_ids=list(affected_product_ids))
        return created_count, updated_count


class UnansweredQuestionUpdateView(StaffRequiredMixin, UpdateView):
    model = UnansweredQuestion
    form_class = UnansweredQuestionForm
    template_name = "consultations/unanswered_form.html"
    success_url = reverse_lazy("consultations:unanswered_list")

    def get_queryset(self):
        return UnansweredQuestion.objects.select_related("customer", "session", "resolved_entry", "product")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["product"] = self.object.product
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["session_messages"] = self.object.session.messages.select_related("product").prefetch_related("knowledge_entries").all()
        context["knowledge_create_url"] = (
            f"{reverse_lazy('knowledge:entry_create')}?unanswered_id={self.object.id}"
            f"&product={self.object.product_id or ''}"
            f"&title={self.object.question[:80]}"
            f"&question={self.object.question}&keywords={self.object.question}"
        )
        context["retest_result"] = self.request.session.pop(f"retest_result_{self.object.id}", None)
        context["message_stats"] = {
            "total": self.object.session.messages.count(),
            "ai": self.object.session.messages.filter(sender=ConsultationMessage.Sender.AI).count(),
            "user": self.object.session.messages.filter(sender=ConsultationMessage.Sender.USER).count(),
        }
        context["page_title"] = "Tinjau Pertanyaan Belum Terjawab"
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        _refresh_quality_from_unanswered(UnansweredQuestion.objects.filter(pk=self.object.pk))
        _log_activity(
            self.request,
            "unanswered_update",
            "unanswered_question",
            self.object.pk,
            f"Perbarui pertanyaan belum terjawab menjadi {self.object.status}",
        )
        return response


class UnansweredQuestionRetestView(StaffRequiredMixin, View):
    def post(self, request, pk, *args, **kwargs):
        unanswered = get_object_or_404(UnansweredQuestion, pk=pk)
        try:
            result = generate_consultation_answer(unanswered.question, product=unanswered.product, request=request)
        except Exception:
            result = {
                "answer": "",
                "confidence": 0,
                "used_rag": True,
                "knowledge_entries": [],
                "source_summary": "",
                "response_state": "fallback",
                "reason": "Terjadi kendala saat menghubungi layanan AI.",
            }
        request.session[f"retest_result_{unanswered.id}"] = {
            "question": unanswered.question,
            "answer": result["answer"],
            "confidence": result["confidence"],
            "source_summary": result["source_summary"],
            "response_state": result["response_state"],
            "reason": result["reason"],
        }
        return redirect("consultations:unanswered_update", pk=unanswered.pk)


class AIResponseReviewListView(StaffRequiredMixin, ListView):
    model = ConsultationMessage
    template_name = "consultations/review_list.html"
    context_object_name = "messages"
    paginate_by = 10

    def post(self, request, *args, **kwargs):
        if request.POST.get("form_type") == "import":
            form = ReviewImportForm(request.POST, request.FILES)
            if not form.is_valid():
                messages.error(request, "File impor evaluasi AI tidak valid.")
                return _redirect_with_query(request, "consultations:review_list")
            created_count, updated_count = self._import_reviews(form.cleaned_data["import_file"])
            messages.success(request, f"Impor evaluasi AI selesai. {created_count} data baru dan {updated_count} data diperbarui.")
            _log_activity(request, "review_import", "ai_review", "", f"Impor evaluasi AI {created_count} baru, {updated_count} diperbarui")
            return _redirect_with_query(request, "consultations:review_list")
        if request.POST.get("form_type") == "bulk_action":
            selected_ids = request.POST.getlist("selected_ids")
            bulk_action = request.POST.get("bulk_action", "").strip()
            if not selected_ids:
                messages.error(request, "Pilih minimal satu jawaban AI untuk aksi massal.")
                return _redirect_with_query(request, "consultations:review_list")
            queryset = ConsultationMessage.objects.filter(
                pk__in=selected_ids,
                sender=ConsultationMessage.Sender.AI,
            )
            if bulk_action in {
                AIResponseReview.Verdict.ACCURATE,
                AIResponseReview.Verdict.NEEDS_REVISION,
                AIResponseReview.Verdict.INCORRECT,
            }:
                affected_count = 0
                for message in queryset:
                    AIResponseReview.objects.update_or_create(
                        message=message,
                        reviewer=request.user,
                        defaults={"verdict": bulk_action, "note": ""},
                    )
                    affected_count += 1
                _refresh_quality_from_messages(queryset)
                messages.success(request, f"{affected_count} evaluasi AI berhasil diperbarui.")
                _log_activity(request, "review_bulk_update", "ai_review", ",".join(selected_ids), f"Evaluasi massal keputusan {bulk_action} untuk {affected_count} jawaban AI")
                return _redirect_with_query(request, "consultations:review_list")
            if bulk_action == "delete_review":
                message_ids = list(queryset.values_list("id", flat=True))
                if request.user.role == "admin" or request.user.is_superuser:
                    deleted_count, _ = AIResponseReview.objects.filter(message__in=queryset).delete()
                else:
                    deleted_count, _ = AIResponseReview.objects.filter(message__in=queryset, reviewer=request.user).delete()
                _refresh_quality_from_messages(ConsultationMessage.objects.filter(pk__in=message_ids))
                messages.success(request, f"{deleted_count} evaluasi AI berhasil dihapus.")
                _log_activity(request, "review_bulk_delete", "ai_review", ",".join(selected_ids), f"Hapus massal {deleted_count} evaluasi AI")
                return _redirect_with_query(request, "consultations:review_list")
            messages.error(request, "Aksi massal evaluasi AI tidak valid.")
            return _redirect_with_query(request, "consultations:review_list")
        return _redirect_with_query(request, "consultations:review_list")

    def get_queryset(self):
        queryset = ConsultationMessage.objects.filter(sender=ConsultationMessage.Sender.AI).select_related("session__user", "session__product", "product")
        verdict = self.request.GET.get("verdict", "").strip()
        review_state = self.request.GET.get("review_state", "").strip()
        response_state = self.request.GET.get("response_state", "").strip()
        product_value = self.request.GET.get("product", "").strip()
        search = self.request.GET.get("q", "").strip()
        if product_value:
            queryset = queryset.filter(product_id=product_value)
        if verdict:
            queryset = queryset.filter(reviews__verdict=verdict)
        if review_state == "pending":
            queryset = queryset.filter(reviews__isnull=True)
        elif review_state == "reviewed":
            queryset = queryset.filter(reviews__isnull=False)
        if response_state:
            queryset = queryset.filter(response_state=response_state)
        if search:
            queryset = queryset.filter(
                Q(content__icontains=search)
                | Q(session__title__icontains=search)
                | Q(session__user__username__icontains=search)
            )
        return queryset.distinct().prefetch_related("reviews")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        accurate_count = AIResponseReview.objects.filter(verdict=AIResponseReview.Verdict.ACCURATE).count()
        needs_revision_count = AIResponseReview.objects.filter(verdict=AIResponseReview.Verdict.NEEDS_REVISION).count()
        incorrect_count = AIResponseReview.objects.filter(verdict=AIResponseReview.Verdict.INCORRECT).count()
        context["filter_values"] = {
            "verdict": self.request.GET.get("verdict", "").strip(),
            "review_state": self.request.GET.get("review_state", "").strip(),
            "response_state": self.request.GET.get("response_state", "").strip(),
            "product": self.request.GET.get("product", "").strip(),
            "q": self.request.GET.get("q", "").strip(),
        }
        product_value = self.request.GET.get("product", "").strip()
        base_queryset = ConsultationMessage.objects.filter(sender=ConsultationMessage.Sender.AI)
        if product_value:
            base_queryset = base_queryset.filter(product_id=product_value)
        context["review_counts"] = {
            "pending": base_queryset.filter(reviews__isnull=True).count(),
            "reviewed": base_queryset.filter(reviews__isnull=False).distinct().count(),
            "fallback": base_queryset.filter(response_state=ConsultationMessage.ResponseState.FALLBACK).count(),
        }
        context["verdict_counts"] = {
            "accurate": accurate_count,
            "needs_revision": needs_revision_count,
            "incorrect": incorrect_count,
        }
        context["products"] = Product.objects.filter(is_active=True).order_by("name")
        context["import_form"] = ReviewImportForm()
        context["query_string"] = _build_query_string(self.request)
        return context

    def _import_reviews(self, uploaded_file):
        rows = _read_import_rows(uploaded_file, "impor evaluasi AI")
        required_columns = {"message_id", "reviewer_username", "product", "verdict", "note"}
        if not rows["fieldnames"]:
            raise ValidationError("Header file evaluasi AI tidak ditemukan.")
        missing_columns = required_columns.difference({field.strip() for field in rows["fieldnames"] if field})
        if missing_columns:
            raise ValidationError(f"Header impor evaluasi AI kurang: {', '.join(sorted(missing_columns))}.")
        valid_verdicts = {choice[0] for choice in AIResponseReview.Verdict.choices}
        created_count = 0
        updated_count = 0
        imported_message_ids = []
        for row in rows["items"]:
            message = ConsultationMessage.objects.filter(
                pk=(row.get("message_id") or "").strip(),
                sender=ConsultationMessage.Sender.AI,
            ).first()
            reviewer = self.request.user.__class__.objects.filter(username=(row.get("reviewer_username") or "").strip()).first()
            product = Product.objects.filter(name=(row.get("product") or "").strip(), is_active=True).first()
            verdict = (row.get("verdict") or "").strip()
            if not message or not reviewer or not product or message.product_id != product.id or verdict not in valid_verdicts:
                continue
            _, created = AIResponseReview.objects.update_or_create(
                message=message,
                reviewer=reviewer,
                defaults={
                    "verdict": verdict,
                    "note": (row.get("note") or "").strip(),
                },
            )
            imported_message_ids.append(message.pk)
            if created:
                created_count += 1
            else:
                updated_count += 1
        _refresh_quality_from_messages(ConsultationMessage.objects.filter(pk__in=imported_message_ids))
        return created_count, updated_count


class AIResponseReviewCreateView(StaffRequiredMixin, CreateView):
    model = AIResponseReview
    form_class = AIResponseReviewForm
    template_name = "consultations/review_form.html"
    success_url = reverse_lazy("consultations:review_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["message_object"] = get_object_or_404(
            ConsultationMessage.objects.select_related("session__user", "product").prefetch_related("knowledge_entries"),
            pk=self.kwargs["message_id"],
        )
        context["session_messages"] = context["message_object"].session.messages.select_related("product").prefetch_related("knowledge_entries").all()
        context["existing_reviews"] = context["message_object"].reviews.select_related("reviewer").all()
        context["page_title"] = "Evaluasi Jawaban AI"
        context["submit_label"] = "Simpan Evaluasi"
        return context

    def form_valid(self, form):
        form.instance.reviewer = self.request.user
        form.instance.message = get_object_or_404(ConsultationMessage, pk=self.kwargs["message_id"])
        response = super().form_valid(form)
        _refresh_quality_from_messages(ConsultationMessage.objects.filter(pk=form.instance.message_id))
        _log_activity(
            self.request,
            "review_create",
            "ai_review",
            self.object.pk,
            f"Membuat evaluasi AI dengan keputusan {self.object.verdict}",
        )
        return response


class AIResponseReviewUpdateView(StaffRequiredMixin, UpdateView):
    model = AIResponseReview
    form_class = AIResponseReviewForm
    template_name = "consultations/review_form.html"
    success_url = reverse_lazy("consultations:review_list")

    def get_queryset(self):
        queryset = AIResponseReview.objects.select_related("message__session__user", "reviewer")
        if self.request.user.role == "ai_trainer" and not self.request.user.is_superuser:
            return queryset.filter(reviewer=self.request.user)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["message_object"] = self.object.message
        context["session_messages"] = self.object.message.session.messages.select_related("product").prefetch_related("knowledge_entries").all()
        context["existing_reviews"] = self.object.message.reviews.select_related("reviewer").all()
        context["page_title"] = "Ubah Evaluasi AI"
        context["submit_label"] = "Perbarui Evaluasi"
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        _refresh_quality_from_messages(ConsultationMessage.objects.filter(pk=self.object.message_id))
        _log_activity(
            self.request,
            "review_update",
            "ai_review",
            self.object.pk,
            f"Memperbarui evaluasi AI menjadi {self.object.verdict}",
        )
        return response


class AIResponseReviewDeleteView(StaffRequiredMixin, DeleteView):
    model = AIResponseReview
    template_name = "consultations/review_confirm_delete.html"
    success_url = reverse_lazy("consultations:review_list")

    def get_queryset(self):
        queryset = AIResponseReview.objects.select_related("message__session__user", "reviewer")
        if self.request.user.role == "ai_trainer" and not self.request.user.is_superuser:
            return queryset.filter(reviewer=self.request.user)
        return queryset

    def form_valid(self, form):
        object_id = self.object.pk
        message_id = self.object.message_id
        response = super().form_valid(form)
        _refresh_quality_from_messages(ConsultationMessage.objects.filter(pk=message_id))
        _log_activity(self.request, "review_delete", "ai_review", object_id, "Menghapus evaluasi AI")
        return response


class UnansweredQuestionDeleteView(StaffRequiredMixin, DeleteView):
    model = UnansweredQuestion
    template_name = "consultations/unanswered_confirm_delete.html"
    success_url = reverse_lazy("consultations:unanswered_list")

    def form_valid(self, form):
        object_id = self.object.pk
        product_id = self.object.product_id
        entry_id = self.object.resolved_entry_id
        response = super().form_valid(form)
        refresh_knowledge_quality(
            product_ids=[product_id] if product_id else None,
            entry_ids=[entry_id] if entry_id else None,
        )
        _log_activity(self.request, "unanswered_delete", "unanswered_question", object_id, "Menghapus pertanyaan belum terjawab")
        return response


class AIResponseReviewExportView(StaffRequiredMixin, ListView):
    model = AIResponseReview

    def get(self, request, *args, **kwargs):
        queryset = AIResponseReview.objects.select_related("message__session__user", "message__session__product", "message__product", "reviewer").order_by("-created_at")
        export_format = request.GET.get("format", "csv").strip().lower()
        search = request.GET.get("q", "").strip()
        review_state = request.GET.get("review_state", "").strip()
        response_state = request.GET.get("response_state", "").strip()
        product = request.GET.get("product", "").strip()
        if review_state == "reviewed":
            queryset = queryset.filter(pk__isnull=False)
        if response_state:
            queryset = queryset.filter(message__response_state=response_state)
        if product:
            queryset = queryset.filter(message__product_id=product)
        if search:
            queryset = queryset.filter(
                Q(message__content__icontains=search)
                | Q(message__session__title__icontains=search)
                | Q(message__session__user__username__icontains=search)
            )
        headers = ["message_id", "product", "session", "customer", "reviewer_username", "verdict", "note", "created_at"]
        rows = [
            [
                review.message_id,
                review.message.product.name if review.message.product else "",
                review.message.session.title,
                review.message.session.user.username,
                review.reviewer.username,
                review.verdict,
                review.note,
                timezone.localtime(review.created_at).strftime("%Y-%m-%d %H:%M:%S"),
            ]
            for review in queryset
        ]
        if export_format == "xlsx":
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(headers)
            for row in rows:
                sheet.append(row)
            buffer = io.BytesIO()
            workbook.save(buffer)
            response = HttpResponse(
                buffer.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = 'attachment; filename="review-ai.xlsx"'
            _log_activity(request, "review_export", "ai_review", "xlsx", "Ekspor evaluasi AI format xlsx")
            return response
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="review-ai.csv"'
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerows(rows)
        _log_activity(request, "review_export", "ai_review", "csv", "Ekspor evaluasi AI format csv")
        return response


class UnansweredQuestionExportView(StaffRequiredMixin, ListView):
    model = UnansweredQuestion

    def get(self, request, *args, **kwargs):
        queryset = UnansweredQuestion.objects.select_related("customer", "resolved_entry", "session", "product").order_by("-created_at")
        status_value = request.GET.get("status", "").strip()
        product = request.GET.get("product", "").strip()
        search = request.GET.get("q", "").strip()
        start_date = parse_date(request.GET.get("start_date", "").strip())
        end_date = parse_date(request.GET.get("end_date", "").strip())
        export_format = request.GET.get("format", "csv").strip().lower()
        if product:
            queryset = queryset.filter(product_id=product)
        if status_value:
            queryset = queryset.filter(status=status_value)
        if search:
            queryset = queryset.filter(
                Q(question__icontains=search)
                | Q(reason__icontains=search)
                | Q(customer__username__icontains=search)
            )
        if start_date:
            queryset = queryset.filter(created_at__date__gte=start_date)
        if end_date:
            queryset = queryset.filter(created_at__date__lte=end_date)
        headers = ["customer", "product", "question", "reason", "status", "resolved_entry", "session", "created_at"]
        rows = [
            [
                item.customer.username,
                item.product.name if item.product else "",
                item.question,
                item.reason,
                item.status,
                item.resolved_entry.title if item.resolved_entry else "",
                item.session.title,
                timezone.localtime(item.created_at).strftime("%Y-%m-%d %H:%M:%S"),
            ]
            for item in queryset
        ]
        if export_format == "xlsx":
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(headers)
            for row in rows:
                sheet.append(row)
            buffer = io.BytesIO()
            workbook.save(buffer)
            response = HttpResponse(
                buffer.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = 'attachment; filename="unanswered-question.xlsx"'
            _log_activity(request, "unanswered_export", "unanswered_question", "xlsx", "Ekspor pertanyaan belum terjawab format xlsx")
            return response
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="unanswered-question.csv"'
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerows(rows)
        _log_activity(request, "unanswered_export", "unanswered_question", "csv", "Ekspor pertanyaan belum terjawab format csv")
        return response
